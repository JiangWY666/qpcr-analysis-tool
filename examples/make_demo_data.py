"""生成一份可以公开分发的 qPCR 演示数据：``examples/demo_qPCR_data.xlsx``。

为什么需要这个脚本
------------------
真实的下机数据属于用户的实验结果，不进版本库。没有它，别人 clone 仓库既跑不了
测试也没法试用软件。这个脚本用 openpyxl 造出一份**结构与 Bio-Rad CFX 导出完全
一致、数值全部虚构**的 96 孔板文件，同时充当两个角色：

- 测试夹具：``tests/test_demo_data.py`` 只依赖它，任何人 clone 下来都能跑通；
- 上手样例：用户下载后直接拖进软件就能走完「读表 → 拆重复 → 选内参 → 导出」。

板面布局（行 = 样本，列块 = 基因，每个区块横向占 3 个连续列）
---------------------------------------------------------
::

            列 01 02 03 | 04 05 06 | 07 08 09 | 10 11 12
    行 A      GAPDH     |   IL6    |   TNF    |   IL1B     样本 Control
    行 B      GAPDH     |   IL6    |   TNF    |   IL1B     样本 Vehicle
    行 C      GAPDH     |   IL6    |   TNF    |   IL1B     样本 Model
    行 D      GAPDH     |   IL6    |   TNF    |   IL1B     样本 Model+Drug
    行 E~H    空孔（凑满 96 孔板）

每个区块内从左到右依次是 1、2、3 号生物学重复，各基因区块的第 j 个孔属于同一只
动物。``replicates.py`` 会把区块识别为「横向排列」并按列号编号，于是 A01+A04+
A07+A10 被配成 ``Control-1``，A02+A05+A08+A11 是 ``Control-2``，依此类推。

数值是怎么编出来的
------------------
全部由固定随机种子生成，反复运行结果完全一致。核心关系是「Cq 每降 1 个循环，
表达量翻一倍」，因此设计倍数 F 对应 ``ΔCt = ΔCt(对照组) − log2(F)``：

- GAPDH 作内参，在所有样本里都稳定在 18 附近，只有微小的上样量漂移；
- IL6 / TNF 在 Model 组相对 Control 上调 4 倍 / 3 倍，Model+Drug 组回落到
  2 倍 / 1.5 倍，体现给药有效；
- IL1B 各组基本不变，充当阴性对照；
- 目标基因的 Cq 建在**该动物自己的 GAPDH Cq** 之上，所以同一号生物学重复的
  内参孔与目标孔共享同一份上样量误差——这正是生物学重复配对的物理来源。

另外故意把 ``B12``（Vehicle / IL1B 的 3 号重复）写成 ``N/A``，用来演示软件对
未检出孔的处理。它落在目标基因而不是内参上：内参孔失效会让整只动物的所有基因
一起退出计算，不适合当演示。

运行方式
--------
::

    .\\.venv\\Scripts\\python.exe examples\\make_demo_data.py

脚本在写完文件后会立刻用 ``read_plate`` 读回来跑一遍完整链路，把识别到的基因、
样本、孔数以及各组的 2^-ΔΔCt 打印出来，生成即自检。
"""

from __future__ import annotations

import math
import random
import statistics
import sys
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:  # 允许从任意工作目录运行本脚本
    sys.path.insert(0, str(ROOT))

from qpcr_tool.analysis import analyze  # noqa: E402
from qpcr_tool.exporter import build_wide_table  # noqa: E402
from qpcr_tool.reader import read_plate  # noqa: E402
from qpcr_tool.replicates import (  # noqa: E402
    group_by_original_sample,
    split_biological_replicates,
)

OUTPUT_FILE = Path(__file__).resolve().parent / "demo_qPCR_data.xlsx"

# --------------------------------------------------------------------------
# 板面布局
# --------------------------------------------------------------------------

SAMPLES = ("Control", "Vehicle", "Model", "Model+Drug")
TARGETS = ("GAPDH", "IL6", "TNF", "IL1B")
REFERENCE_TARGET = "GAPDH"
CONTROL_SAMPLE = "Control"
REPLICATES = 3

# 样本 -> 板上的行号；基因 -> 该基因区块的起始列号（区块向右占 3 列）
SAMPLE_ROW = {"Control": "A", "Vehicle": "B", "Model": "C", "Model+Drug": "D"}
TARGET_FIRST_COLUMN = {"GAPDH": 1, "IL6": 4, "TNF": 7, "IL1B": 10}

PLATE_ROWS = "ABCDEFGH"
PLATE_COLUMNS = 12

# 未检出的孔：Vehicle / IL1B 的 3 号重复。写成 CFX 里常见的 "N/A"。
UNDETERMINED_WELL = "B12"
UNDETERMINED_TEXT = "N/A"

# --------------------------------------------------------------------------
# 数值设计
# --------------------------------------------------------------------------

RANDOM_SEED = 20260605

# 各样本的内参基线 Cq。样本之间只有 ±0.06 的上样量漂移，内参该有的样子。
GAPDH_BASELINE = {
    "Control": 18.00,
    "Vehicle": 18.06,
    "Model": 17.94,
    "Model+Drug": 18.02,
}

# 对照组里各目标基因相对内参的 ΔCt（Cq_target − Cq_GAPDH），决定基因的丰度。
CONTROL_DCT = {"IL6": 6.20, "TNF": 5.40, "IL1B": 7.10}

# 设计倍数：相对 Control 组的 2^-ΔΔCt。IL1B 三组都在 1 附近，是阴性对照。
FOLD_CHANGES = {
    "IL6": {"Control": 1.00, "Vehicle": 1.05, "Model": 4.00, "Model+Drug": 2.00},
    "TNF": {"Control": 1.00, "Vehicle": 0.95, "Model": 3.00, "Model+Drug": 1.50},
    "IL1B": {"Control": 1.00, "Vehicle": 1.04, "Model": 1.10, "Model+Drug": 0.96},
}

# 同一样本 3 只动物之间的上样量差异，内参与目标基因共享
ANIMAL_SD_RANGE = (0.08, 0.14)
# 同一只动物同一基因的测量噪声，叠在动物偏移之上
RESIDUAL_SD_RANGE = (0.04, 0.10)

# 写进表格的 Cq 小数位。真实 CFX 会导出十几位，四位既够用又便于人工核对。
CQ_DECIMALS = 4

# --------------------------------------------------------------------------
# CFX 导出格式
# --------------------------------------------------------------------------

DATA_SHEET_NAME = "0"
RUN_INFO_SHEET_NAME = "Run Information"

# 表头从 B 列开始，A 列留空——这是 Bio-Rad CFX 导出的固有形状，一并模仿。
HEADERS = (
    "Well", "Fluor", "Target", "Content", "Sample", "Biological Set Name",
    "Cq", "Cq Mean", "Cq Std. Dev", "Starting Quantity (SQ)",
    "Log Starting Quantity", "SQ Mean", "SQ Std. Dev", "Set Point", "Well Note",
)
FIRST_HEADER_COLUMN = 2

FLUOR = "SYBR"
CONTENT = "Unkn"
SET_POINT = 55

RUN_INFORMATION = (
    ("File Name", "demo_qPCR_data.pcrd"),
    ("Created By User", "demo"),
    ("Notes", "虚构的演示数据，不含任何真实实验结果"),
    ("ID", ""),
    ("Run Started", "01/15/2026 09:30:00 UTC"),
    ("Run Ended", "01/15/2026 10:38:00 UTC"),
    ("Sample Vol", 20),
    ("Lid Temp", 105),
    ("Protocol File Name", "qPCR-demo.prcl"),
    ("Plate Setup File Name", "Quick Plate_96 wells_SYBR Only.pltd"),
    ("Base Serial Number", "DEMO0001"),
    ("Optical Head Serial Number", "DEMO-HEAD-01"),
    ("Run Type", "Quantification"),
)


# --------------------------------------------------------------------------
# 生成数值
# --------------------------------------------------------------------------


def shaped_offsets(rng: random.Random, count: int, sd: float) -> list[float]:
    """抽 count 个正态偏移，再平移缩放成「均值恰为 0、样本标准差恰为 sd」。

    这么做有两个目的：一是离散度可控，不会因为随机种子的运气抽出一个大得离谱的
    标准差，内参「复孔 SD < 0.2」这类约束才有保证；二是均值恰为 0，设计好的倍数
    关系能被下游精确复现，测试敢用较紧的容差去卡。组内各孔仍然彼此不同，看上去
    依旧是一批有测量误差的真实读数。
    """
    raw = [rng.gauss(0.0, 1.0) for _ in range(count)]
    mean = statistics.fmean(raw)
    centered = [value - mean for value in raw]
    spread = statistics.stdev(centered)
    if spread == 0:  # 理论上抽不到，兜一手避免除零
        return [0.0] * count
    return [value * sd / spread for value in centered]


def build_cq_table() -> dict[tuple[str, str], list[float]]:
    """算出每个 (样本, 基因) 的 3 个 Cq，返回 {(样本, 基因): [复孔1, 复孔2, 复孔3]}。

    列表下标 j 对应第 j+1 号生物学重复。目标基因的 Cq 由「该动物的 GAPDH Cq
    + 设计 ΔCt + 测量噪声」得到，所以内参孔和目标孔天然共享同一份上样量误差。
    """
    rng = random.Random(RANDOM_SEED)
    table: dict[tuple[str, str], list[float]] = {}
    for sample in SAMPLES:
        animal = shaped_offsets(rng, REPLICATES, rng.uniform(*ANIMAL_SD_RANGE))
        table[(sample, REFERENCE_TARGET)] = [
            round(GAPDH_BASELINE[sample] + offset, CQ_DECIMALS) for offset in animal
        ]
        for target in TARGETS:
            if target == REFERENCE_TARGET:
                continue
            # Cq 每降 1 个循环表达量翻倍，所以倍数 F 对应 ΔCt 下降 log2(F)
            dct = CONTROL_DCT[target] - math.log2(FOLD_CHANGES[target][sample])
            residual = shaped_offsets(rng, REPLICATES, rng.uniform(*RESIDUAL_SD_RANGE))
            table[(sample, target)] = [
                round(
                    GAPDH_BASELINE[sample] + animal[index] + dct + residual[index],
                    CQ_DECIMALS,
                )
                for index in range(REPLICATES)
            ]
    return table


def well_name(row: str, column: int) -> str:
    """(行字母, 列号) -> CFX 风格的孔位字符串，列号补零成两位。"""
    return f"{row}{column:02d}"


def build_layout() -> dict[str, tuple[str, str, int]]:
    """孔位 -> (样本, 基因, 生物学重复编号)，只含上了样的 48 个孔。"""
    layout: dict[str, tuple[str, str, int]] = {}
    for sample in SAMPLES:
        row = SAMPLE_ROW[sample]
        for target in TARGETS:
            first = TARGET_FIRST_COLUMN[target]
            for index in range(REPLICATES):
                layout[well_name(row, first + index)] = (sample, target, index + 1)
    return layout


def all_wells() -> list[str]:
    """整块 96 孔板的孔位，按 CFX 的行优先顺序：A01…A12、B01…B12、…、H12。"""
    return [
        well_name(row, column)
        for row in PLATE_ROWS
        for column in range(1, PLATE_COLUMNS + 1)
    ]


# --------------------------------------------------------------------------
# 写文件
# --------------------------------------------------------------------------


def _row_values(
    well: str,
    layout: dict[str, tuple[str, str, int]],
    cq_table: dict[tuple[str, str], list[float]],
) -> list[object]:
    """一个孔在数据 sheet 里的 15 个单元格（从 B 列开始），顺序同 HEADERS。"""
    assignment = layout.get(well)
    if assignment is None:
        # 空孔：CFX 照样会为它写一行，Target 与 Sample 留空，read_plate 会跳过
        return [well, FLUOR, "", CONTENT, "", "", None, 0, 0, None, None, 0, 0,
                SET_POINT, ""]

    sample, target, index = assignment
    cq: object = (
        UNDETERMINED_TEXT
        if well == UNDETERMINED_WELL
        else cq_table[(sample, target)][index - 1]
    )
    # 真实文件里 Cq Mean 就等于该孔自身的 Cq、Std. Dev 为 0（CFX 未设复孔编号）
    return [well, FLUOR, target, CONTENT, sample, "", cq, cq, 0, None, None, None,
            0, SET_POINT, ""]


def write_workbook(path: Path, cq_table: dict[tuple[str, str], list[float]]) -> None:
    """把 96 孔写成 CFX 风格的双 sheet xlsx。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = DATA_SHEET_NAME

    for offset, title in enumerate(HEADERS):
        sheet.cell(row=1, column=FIRST_HEADER_COLUMN + offset, value=title)

    layout = build_layout()
    for row_number, well in enumerate(all_wells(), start=2):
        for offset, value in enumerate(_row_values(well, layout, cq_table)):
            sheet.cell(row=row_number, column=FIRST_HEADER_COLUMN + offset, value=value)

    info_sheet = workbook.create_sheet(RUN_INFO_SHEET_NAME)
    for row_number, (key, value) in enumerate(RUN_INFORMATION, start=1):
        info_sheet.cell(row=row_number, column=1, value=key)
        info_sheet.cell(row=row_number, column=2, value=value)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(path))
    workbook.close()


# --------------------------------------------------------------------------
# 生成后自检
# --------------------------------------------------------------------------


def verify(path: Path) -> None:
    """读回刚写出的文件，跑一遍完整链路并打印关键数值。

    链路与软件里的默认口径一致：拆生物学重复 → 按原样本名分组 → GAPDH 归一化、
    Control 作对照 → 宽表。任何一步出问题都会直接抛异常，生成即自检。
    """
    plate = read_plate(str(path))
    print(f"文件      : {path}")
    print(f"数据 sheet: {plate.sheet_name}")
    print(f"基因      : {plate.targets}")
    print(f"样本      : {plate.samples}")
    print(
        f"孔数      : {len(plate.wells)}"
        f"（有效 {sum(1 for w in plate.wells if w.valid)}，"
        f"无效 {len(plate.invalid_wells)}）"
    )
    for well in plate.invalid_wells:
        print(f"  未检出孔: {well.well}  {well.target} / {well.sample}  "
              f"Cq 原文 {well.cq_text!r}")
    for warning in plate.warnings:
        print(f"[读取警告] {warning}")

    print("\n内参复孔一致性（GAPDH 三个孔的 Cq）:")
    for sample in SAMPLES:
        values = [w.cq for w in plate.wells
                  if w.target == REFERENCE_TARGET and w.original_sample == sample]
        print(f"  {sample:<12s} {[f'{v:.4f}' for v in values]}  "
              f"均值 {statistics.fmean(values):.4f}  "
              f"SD {statistics.stdev(values):.4f}")

    report = split_biological_replicates(plate)
    groups = group_by_original_sample(plate)
    print(f"\n拆分后的虚拟样本: {plate.samples}")
    print(f"区块方向        : {sorted(set(report.orientations.values()))}")
    for warning in report.warnings:
        print(f"[拆分提示] {warning}")

    result = analyze(plate.wells, groups, [REFERENCE_TARGET], CONTROL_SAMPLE)
    for warning in result.warnings:
        print(f"[分析警告] {warning}")

    print(f"\n2^-ΔΔCt（内参 {REFERENCE_TARGET}，对照组 {CONTROL_SAMPLE}）:")
    print(f"  {'基因':<6s}{'分组':<14s}{'n':>3s}{'RQ 均值':>11s}"
          f"{'RQ 标准差':>12s}{'设计倍数':>10s}")
    for target in result.targets:
        for group in result.groups:
            stat = result.stat(target, group)
            if stat is None or stat.mean is None:
                continue
            sd = f"{stat.sd:.4f}" if stat.sd is not None else "-"
            print(f"  {target:<6s}{group:<14s}{stat.n:>3d}{stat.mean:>11.4f}"
                  f"{sd:>12s}{FOLD_CHANGES[target][group]:>10.2f}")

    table = build_wide_table(result)
    print(f"\n宽表: {len(table.columns)} 列 × {len(table.rows)} 行"
          f"（列 = {len(result.targets)} 基因 × {len(result.groups)} 分组）")


def main() -> None:
    write_workbook(OUTPUT_FILE, build_cq_table())
    verify(OUTPUT_FILE)


if __name__ == "__main__":
    main()
