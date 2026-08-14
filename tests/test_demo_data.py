"""只依赖仓库自带演示数据的端到端测试。

任何人 clone 下来、CI 环境里都能跑。真实下机数据的精确断言在
``test_pipeline.py`` / ``test_replicates.py`` / ``test_gui_smoke.py`` 里，
那些文件缺失时会优雅跳过，不在这里重复。
"""

from __future__ import annotations

import os
import statistics
import tempfile
import unittest

try:
    from ._fixtures import DEMO_FILE, requires_demo_file
except ImportError:
    from _fixtures import DEMO_FILE, requires_demo_file

from openpyxl import load_workbook

from qpcr_tool.analysis import analyze
from qpcr_tool.exporter import build_wide_table, export_excel
from qpcr_tool.reader import read_plate
from qpcr_tool.replicates import (
    build_pairing_preview,
    group_by_original_sample,
    split_biological_replicates,
)

TARGETS = ["GAPDH", "IL6", "TNF", "IL1B"]
SAMPLES = ["Control", "Vehicle", "Model", "Model+Drug"]
REFERENCE = ["GAPDH"]
CONTROL = "Control"

# 演示数据里设计的相对表达量（相对 Control）。容差留足生成噪声。
EXPECTED_FOLD = {
    "IL6": {"Model": 4.0, "Model+Drug": 2.0},
    "TNF": {"Model": 3.0, "Model+Drug": 1.5},
}
FOLD_TOLERANCE = 0.15  # ±15%

# Vehicle / IL1B 的 3 号重复故意写成 N/A
UNDETERMINED_WELL = "B12"


@requires_demo_file
class TestDemoDataPipeline(unittest.TestCase):
    """演示数据：读表 → 拆分 → 计算 → 宽表 → 导出。"""

    @classmethod
    def setUpClass(cls) -> None:
        cls.plate = read_plate(str(DEMO_FILE))
        # 拆分会就地改写 sample 名，先把原始识别结果存下来给读表断言用
        cls.raw_targets = list(cls.plate.targets)
        cls.raw_samples = list(cls.plate.samples)
        cls.split_report = split_biological_replicates(cls.plate)
        cls.groups = group_by_original_sample(cls.plate)
        cls.result = analyze(cls.plate.wells, cls.groups, REFERENCE, CONTROL)
        cls.wide = build_wide_table(cls.result)

    def test_01_read_plate(self) -> None:
        self.assertEqual(self.raw_targets, TARGETS)
        self.assertEqual(self.raw_samples, SAMPLES)
        self.assertEqual(len(self.plate.wells), 48)
        self.assertEqual(len(self.plate.invalid_wells), 1)
        invalid = self.plate.invalid_wells[0]
        self.assertEqual(invalid.well, UNDETERMINED_WELL)
        self.assertEqual(invalid.target, "IL1B")
        # 拆分后 sample 变成 Vehicle-3，原名在 original_sample
        self.assertEqual(invalid.original_sample or invalid.sample, "Vehicle")
        self.assertFalse(invalid.included)

    def test_02_biological_split_and_pairing(self) -> None:
        self.assertEqual(self.split_report.split_samples, SAMPLES)
        self.assertFalse(self.split_report.warnings)

        # 12 个虚拟样本，每个恰好 4 个基因孔
        virtual = sorted({w.sample for w in self.plate.wells})
        self.assertEqual(len(virtual), 12)
        for name in virtual:
            wells = [w for w in self.plate.wells if w.sample == name]
            self.assertEqual(len(wells), 4, name)
            self.assertEqual(sorted(w.target for w in wells), sorted(TARGETS))

        # Control-1 = A01(GAPDH) + A04(IL6) + A07(TNF) + A10(IL1B)
        by_name = {
            w.sample: {x.target: x.well for x in self.plate.wells if x.sample == w.sample}
            for w in self.plate.wells
        }
        self.assertEqual(
            by_name["Control-1"],
            {"GAPDH": "A01", "IL6": "A04", "TNF": "A07", "IL1B": "A10"},
        )
        self.assertEqual(
            by_name["Control-3"],
            {"GAPDH": "A03", "IL6": "A06", "TNF": "A09", "IL1B": "A12"},
        )
        self.assertEqual(
            by_name["Model-2"],
            {"GAPDH": "C02", "IL6": "C05", "TNF": "C08", "IL1B": "C11"},
        )

    def test_03_pairing_preview(self) -> None:
        previews = {p.sample: p for p in build_pairing_preview(self.plate)}
        self.assertEqual(sorted(previews), sorted(SAMPLES))
        control = previews["Control"]
        self.assertIn("横", control.orientation_note)
        self.assertEqual(len(control.rows), 3)
        rep1 = dict(control.rows)[1]
        self.assertEqual(rep1["GAPDH"], "A01")
        self.assertEqual(rep1["IL6"], "A04")
        self.assertEqual(rep1["TNF"], "A07")
        self.assertEqual(rep1["IL1B"], "A10")

    def test_04_control_rq_near_one(self) -> None:
        self.assertEqual(self.result.targets, ["IL6", "TNF", "IL1B"])
        for target in self.result.targets:
            values = self.result.values(target, CONTROL)
            self.assertGreaterEqual(len(values), 2, target)
            self.assertAlmostEqual(statistics.fmean(values), 1.0, delta=0.05, msg=target)
            # 几何平均在数学上应精确等于 1
            self.assertAlmostEqual(
                statistics.geometric_mean(values), 1.0, places=6, msg=target
            )

    def test_05_designed_fold_changes(self) -> None:
        for target, groups in EXPECTED_FOLD.items():
            for group, expected in groups.items():
                values = self.result.values(target, group)
                self.assertTrue(values, f"{target}/{group} 无数据")
                mean = statistics.fmean(values)
                lo, hi = expected * (1 - FOLD_TOLERANCE), expected * (1 + FOLD_TOLERANCE)
                self.assertGreaterEqual(mean, lo, f"{target}/{group}={mean:.3f}")
                self.assertLessEqual(mean, hi, f"{target}/{group}={mean:.3f}")

        # IL1B 是阴性对照，各组应在 1 附近
        for group in SAMPLES:
            values = self.result.values("IL1B", group)
            if not values:
                continue
            mean = statistics.fmean(values)
            self.assertGreaterEqual(mean, 0.7, f"IL1B/{group}")
            self.assertLessEqual(mean, 1.4, f"IL1B/{group}")

    def test_06_undetermined_well_excluded(self) -> None:
        # B12 无效，Vehicle / IL1B 只有 2 个 RQ
        self.assertEqual(len(self.result.values("IL1B", "Vehicle")), 2)
        usable = [
            w for w in self.plate.wells
            if w.well == UNDETERMINED_WELL and w.target == "IL1B"
        ]
        self.assertEqual(len(usable), 1)
        self.assertFalse(usable[0].usable)

    def test_07_wide_table_shape(self) -> None:
        # 3 基因 × 4 组 = 12 列；最长组 3 行（IL1B/Vehicle 只有 2，其它 3）
        self.assertEqual(len(self.wide.columns), 12)
        self.assertEqual(len(self.wide.rows), 3)
        spans = self.wide.merge_spans()
        self.assertEqual(len(spans), 3)
        for start, end, gene in spans:
            self.assertEqual(end - start + 1, 4, gene)

    def test_08_export_excel(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "demo_out.xlsx")
            export_excel(self.result, self.plate, path, self.split_report)
            wb = load_workbook(path, data_only=True)
            try:
                self.assertIn("Prism", wb.sheetnames)
                self.assertIn("明细QC", wb.sheetnames)
                prism = wb["Prism"]
                # 第 2 行应是组名，Control 出现在第一块
                self.assertEqual(prism.cell(2, 1).value, "Control")
                qc = wb["明细QC"]
                # 分析参数里应能看到生物学重复模式
                texts = [
                    str(cell.value)
                    for row in qc.iter_rows(values_only=False)
                    for cell in row
                    if cell.value is not None
                ]
                joined = " ".join(texts)
                self.assertIn("生物学重复", joined)
            finally:
                wb.close()

    def test_09_one_to_one_reference_pairing(self) -> None:
        """拆分后目标孔的内参基线必须等于同号重复的 GAPDH Cq 原值。"""
        gapdh = {
            w.sample: w.cq
            for w in self.plate.wells
            if w.target == "GAPDH" and w.usable
        }
        for item in self.result.well_results:
            if item.sample not in gapdh:
                continue
            self.assertAlmostEqual(
                item.ref_mean, gapdh[item.sample], places=12,
                msg=f"{item.well}/{item.target}/{item.sample}",
            )


if __name__ == "__main__":
    unittest.main()
