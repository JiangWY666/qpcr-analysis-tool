"""GUI 冒烟测试：在无显示环境下走完「加载 → 计算 → 复制」主流程。

必须在导入 PySide6 之前把 Qt 平台插件切成 offscreen，否则无桌面会话时会直接崩。

除了「窗口能建起来」这条，其余用例都要加载真实下机数据。那份文件不进版本库，
缺失时它们会跳过，只留下不依赖数据的窗口初始化冒烟。
"""

from __future__ import annotations

import os
import tempfile
import unittest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

try:  # 兼容 discover -s tests、python -m unittest tests.xxx、直接运行本文件三种方式
    from ._fixtures import SAMPLE_FILE, requires_sample_file
except ImportError:
    from _fixtures import SAMPLE_FILE, requires_sample_file

from openpyxl import Workbook, load_workbook  # noqa: E402
from PySide6.QtWidgets import QApplication  # noqa: E402 - 必须在设置环境变量之后

from qpcr_tool.gui import MainWindow  # noqa: E402

REFERENCE_GENE = "GAPDH"
CONTROL_GROUP = "CT"
EXPECTED_TARGETS = 4
EXPECTED_GROUPS = 4
EXPECTED_WELLS = 48
EXPECTED_WIDE_COLUMNS = 12  # 3 个目标基因 × 4 个分组
EXPECTED_WIDE_ROWS = 3      # 每个 (基因, 组) 有 3 个复孔

ORIGINAL_SAMPLES = ["CT", "PBS BALF", "LPS BALF", "LPS BALF CIT013"]
VIRTUAL_SAMPLES = [f"{s}-{i}" for s in ORIGINAL_SAMPLES for i in (1, 2, 3)]


def sample_file() -> str:
    """真实下机样本文件的路径。定位逻辑集中在 tests/_fixtures.py。"""
    return str(SAMPLE_FILE)


def write_plate_file(
    directory: str, name: str, samples: list[str], targets: list[str]
) -> str:
    """写一份最小可读的下机表：每个 (基因, 样本) 两个横向排列的孔。

    用于不依赖真实数据、需要特定样本命名的用例。
    """
    book = Workbook()
    sheet = book.active
    sheet.append(["Well", "Target", "Sample", "Cq"])
    for gene_index, target in enumerate(targets):
        for offset in (0, 1):
            row_letter = chr(ord("A") + gene_index * 2 + offset)
            for column, sample in enumerate(samples, start=1):
                sheet.append(
                    [f"{row_letter}{column:02d}", target, sample, 20.0 + column]
                )
    path = os.path.join(directory, name)
    book.save(path)
    book.close()
    return path


class GuiSmokeTest(unittest.TestCase):
    """整套流程都通过公开方法驱动，不弹任何对话框。"""

    app: QApplication

    @classmethod
    def setUpClass(cls) -> None:
        cls.app = QApplication.instance() or QApplication([])

    def setUp(self) -> None:
        self.window = MainWindow()
        self._tmp = tempfile.TemporaryDirectory()
        self.tmpdir = self._tmp.name

    def tearDown(self) -> None:
        self.window.close()
        self.window.deleteLater()
        self.app.processEvents()
        self._tmp.cleanup()

    def test_01_window_created(self) -> None:
        self.assertEqual(self.window.windowTitle(), "qPCR 分析工具 v1.0")
        self.assertFalse(self.window.calc_btn.isEnabled())
        self.assertFalse(self.window.export_btn.isEnabled())

    @requires_sample_file
    def test_02_load_sample_file(self) -> None:
        plate = self.window.load_file(sample_file())
        self.assertEqual(len(plate.wells), EXPECTED_WELLS)
        self.assertEqual(self.window.target_list.count(), EXPECTED_TARGETS)
        self.assertEqual(self.window.group_table.rowCount(), EXPECTED_GROUPS)
        self.assertEqual(self.window.well_table.rowCount(), EXPECTED_WELLS)
        self.assertTrue(self.window.calc_btn.isEnabled())
        # GAPDH 应被自动识别成内参，CT 应被猜成对照组
        self.assertIn(REFERENCE_GENE, self.window.reference_targets())
        self.assertEqual(self.window.control_group_name(), CONTROL_GROUP)

    @requires_sample_file
    def test_03_run_analysis(self) -> None:
        self.window.load_file(sample_file())
        self.window.set_reference_targets([REFERENCE_GENE])
        self.window.set_control_group(CONTROL_GROUP)
        result = self.window.run_analysis()

        self.assertEqual(result.control_group, CONTROL_GROUP)
        self.assertEqual(len(result.targets), EXPECTED_TARGETS - 1)
        self.assertEqual(self.window.wide_view.columnCount(), EXPECTED_WIDE_COLUMNS)
        self.assertEqual(len(self.window.wide.rows), EXPECTED_WIDE_ROWS)
        self.assertEqual(
            self.window.wide_view.rowCount(),
            MainWindow.WIDE_HEADER_ROWS + EXPECTED_WIDE_ROWS,
        )
        self.assertEqual(
            self.window.summary_table.rowCount(),
            (EXPECTED_TARGETS - 1) * EXPECTED_GROUPS,
        )
        self.assertTrue(self.window.export_btn.isEnabled())

    @requires_sample_file
    def test_04_copy_to_clipboard(self) -> None:
        self.window.load_file(sample_file())
        self.window.set_reference_targets([REFERENCE_GENE])
        self.window.set_control_group(CONTROL_GROUP)
        self.window.run_analysis()

        columns, rows = self.window.copy_to_clipboard(True)
        self.assertEqual((columns, rows), (EXPECTED_WIDE_COLUMNS, EXPECTED_WIDE_ROWS))
        self.assertEqual(
            self.window.copy_to_clipboard(False),
            (EXPECTED_WIDE_COLUMNS, EXPECTED_WIDE_ROWS),
        )

    def test_05a_merge_checkbox_works_while_split_is_on(self) -> None:
        """回归：PBS1/PBS2 这种「不同名的生物学重复」要能靠合并选项并成一组。

        这两个开关是正交的——拆分处理同名多孔，合并处理不同名但同属一个实验组，
        一块板上经常同时出现。曾经把合并选项置灰，用户点了没反应。
        """
        path = write_plate_file(
            self.tmpdir,
            "命名重复.xlsx",
            samples=["PBS1", "G43-L1", "PBS2", "G43-L2"],
            targets=["GAPDH", "IL6"],
        )
        self.window.load_file(path)

        self.assertTrue(self.window.split_check.isChecked())
        self.assertTrue(self.window.merge_check.isEnabled())
        self.assertEqual(
            [g.name for g in self.window.groups],
            ["PBS1", "G43-L1", "PBS2", "G43-L2"],
        )

        self.window.merge_check.setChecked(True)
        self.assertEqual([g.name for g in self.window.groups], ["PBS", "G43-L"])
        self.assertEqual(self.window.group_table.rowCount(), 2)
        # 组内仍是拆分后的虚拟样本名，配对关系不受并组影响
        self.assertEqual(
            self.window.groups[0].samples, ["PBS1-1", "PBS1-2", "PBS2-1", "PBS2-2"]
        )

        self.window.merge_check.setChecked(False)
        self.assertEqual(len(self.window.groups), 4)

    @requires_sample_file
    def test_05_biological_replicate_defaults(self) -> None:
        """默认勾选生物学重复：分组回到原始样本名，孔位表显示虚拟样本与重复号。"""
        self.assertTrue(self.window.split_check.isChecked())
        self.window.load_file(sample_file())

        self.assertEqual(self.window.fill_direction(), "auto")
        self.assertTrue(self.window.split_report.enabled)
        self.assertEqual(len(self.window.split_report.split_samples), EXPECTED_GROUPS)
        self.assertEqual(self.window.plate.samples, VIRTUAL_SAMPLES)
        self.assertEqual([g.name for g in self.window.groups], ORIGINAL_SAMPLES)
        self.assertEqual(self.window.group_table.rowCount(), EXPECTED_GROUPS)
        # 回归：拆分开着时「合并末尾编号」也必须可点，PBS1/PBS2 这种命名要靠它并组
        self.assertTrue(self.window.merge_check.isEnabled())
        self.assertFalse(self.window.merge_check.isChecked())
        self.assertTrue(self.window.direction_combo.isEnabled())
        # 无显示环境下 isVisible 恒为假，只能看有没有被显式隐藏
        self.assertTrue(self.window.alert_banner.isHidden())
        self.assertEqual(self.window.collect_alerts(), [])

        header = [
            self.window.well_table.horizontalHeaderItem(i).text()
            for i in range(self.window.well_table.columnCount())
        ]
        self.assertEqual(header[3:5], ["样本", "重复"])
        first = [self.window.well_table.item(0, c).text() for c in (1, 3, 4)]
        self.assertEqual(first, ["A01", "CT-1", "1"])

    @requires_sample_file
    def test_06_pairing_preview_tab(self) -> None:
        """配对预览页签夹在孔位数据和分析结果之间，内容与实际配对一致。"""
        self.window.load_file(sample_file())
        titles = [self.window.tabs.tabText(i) for i in range(self.window.tabs.count())]
        self.assertEqual(titles, ["孔位数据", "配对预览", "分析结果"])

        table = self.window.pairing_table
        columns = [
            table.horizontalHeaderItem(i).text() for i in range(table.columnCount())
        ]
        self.assertEqual(columns, ["重复", "GAPDH", "IL1b", "IL6", "TNFa"])
        # 4 个样本，各一行小节标题加三行重复
        self.assertEqual(table.rowCount(), EXPECTED_GROUPS * 4)
        self.assertIn("CT", table.item(0, 0).text())
        self.assertIn("横向排列", table.item(0, 0).text())
        self.assertEqual(
            [table.item(1, c).text() for c in range(5)],
            ["重复 1", "A01", "A04", "E01", "E04"],
        )
        self.assertEqual(
            [table.item(3, c).text() for c in range(5)],
            ["重复 3", "A03", "A06", "E03", "E06"],
        )
        self.assertEqual([p.sample for p in self.window.previews], ORIGINAL_SAMPLES)

    @requires_sample_file
    def test_07_disable_split_falls_back_to_averaging(self) -> None:
        """取消勾选后能重新算出结果，且数值回到内参取平均的口径。"""
        self.window.load_file(sample_file())
        self.window.set_reference_targets([REFERENCE_GENE])
        split_result = self.window.run_analysis()
        split_rq = {(r.target, r.well): r.rq for r in split_result.well_results}

        self.window.set_split_enabled(False)
        self.assertIsNone(self.window.split_report)
        self.assertEqual(self.window.plate.samples, ORIGINAL_SAMPLES)
        self.assertTrue(self.window.merge_check.isEnabled())
        self.assertFalse(self.window.direction_combo.isEnabled())
        self.assertEqual([g.name for g in self.window.groups], ORIGINAL_SAMPLES)

        self.window.set_control_group(CONTROL_GROUP)
        plain_result = self.window.run_analysis()
        plain_rq = {(r.target, r.well): r.rq for r in plain_result.well_results}
        self.assertEqual(set(plain_rq), set(split_rq))
        self.assertEqual(len(self.window.wide.rows), EXPECTED_WIDE_ROWS)

        # 未拆分时 A04 的内参基线就是 CT 三个 GAPDH 孔的均值
        cq = {w.well: w.cq for w in self.window.plate.wells}
        expected = sum(cq[w] for w in ("A01", "A02", "A03")) / 3
        a04 = next(r for r in plain_result.well_results if r.well == "A04")
        self.assertAlmostEqual(a04.ref_mean, expected, places=12)
        self.assertNotAlmostEqual(split_rq[("IL1b", "A04")], plain_rq[("IL1b", "A04")], places=6)

    @requires_sample_file
    def test_08_direction_switch_rebuilds_pairing(self) -> None:
        """切换孔位排列方向要重做拆分；真实板子是一维的，结果不应改变。"""
        self.window.load_file(sample_file())
        before = [(w.well, w.sample, w.replicate_index) for w in self.window.plate.wells]
        self.window.set_fill_direction("column")
        self.assertEqual(self.window.fill_direction(), "column")
        self.assertEqual(self.window.split_report.fill_direction, "column")
        after = [(w.well, w.sample, w.replicate_index) for w in self.window.plate.wells]
        self.assertEqual(before, after)
        self.assertEqual(self.window.pairing_table.rowCount(), EXPECTED_GROUPS * 4)

    @requires_sample_file
    def test_09_alert_banner_shows_dropout(self) -> None:
        """内参孔失效导致整只动物出局时，顶部警告条必须出现。"""
        self.window.load_file(sample_file())
        self.window.set_reference_targets([REFERENCE_GENE])
        for well in self.window.plate.wells:
            if well.well == "A02":  # CT-2 唯一的内参孔
                well.included = False
        self.window.run_analysis()

        self.assertFalse(self.window.alert_banner.isHidden())
        text = self.window.alert_banner.text()
        self.assertIn("CT 的第 2 号生物学重复", text)
        self.assertIn("整只动物的所有基因", text)
        self.assertEqual(len(self.window.collect_alerts()), 2)

    @requires_sample_file
    def test_10_export_keeps_the_mode_that_produced_the_numbers(self) -> None:
        """算完再改开关时，导出的报告要跟数值对得上，明细也不能重复列孔。"""
        self.window.load_file(sample_file())
        self.window.set_reference_targets([REFERENCE_GENE])
        self.window.run_analysis()
        self.window.set_split_enabled(False)  # 结果仍是拆分模式算出来的

        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "导出.xlsx")
            self.window.export_to(path)
            workbook = load_workbook(path)
            try:
                rows = list(workbook["明细QC"].iter_rows(values_only=True))
            finally:
                workbook.close()

        params = {row[0]: row[1] for row in rows if isinstance(row[0], str)}
        self.assertEqual(params["复孔处理模式"], "生物学重复配对（同名样本已拆分）")

        header_idx = [row[0] for row in rows].index("逐孔明细") + 1
        detail = rows[header_idx + 1: header_idx + 1 + EXPECTED_WELLS]
        # 每个孔只能出现一次：样本名变了也不能让同一个孔既算「进了结果」又算「被排除」
        self.assertEqual(len({(r[0], r[1]) for r in detail}), EXPECTED_WELLS)
        self.assertEqual(len([r for r in detail if r[5] == "是"]), 36)
        # 进了结果的孔仍带着算它时用的虚拟样本名，原样本名列能对回下机表
        il1b = [r for r in detail if r[1] == "IL1b"]
        self.assertEqual({r[2] for r in il1b}, {f"CT-{i}" for i in (1, 2, 3)}
                         | {f"{s}-{i}" for s in ORIGINAL_SAMPLES[1:] for i in (1, 2, 3)})
        self.assertEqual({r[10] for r in detail}, set(ORIGINAL_SAMPLES))


if __name__ == "__main__":
    unittest.main(verbosity=2)
