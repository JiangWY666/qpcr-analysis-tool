"""端到端流水线测试：读表 → 分组 → 2^-ΔΔCt 计算 → 宽表 → Excel 导出。

本文件里凡是读真实下机数据的用例都挂了 ``@requires_sample_file``：那份文件是用户
的实验结果，不进版本库，缺失时这些用例会整体跳过而不是报错。不依赖它的纯逻辑
用例（如样本名剥编号）任何环境下都照常执行。只用演示数据的完整链路测试在
``tests/test_demo_data.py``。

用标准库 unittest，从项目根目录运行：
    .\\.venv\\Scripts\\python.exe -m unittest discover -s tests -v
"""

from __future__ import annotations

import math
import os
import tempfile
import unittest

try:  # 兼容 discover -s tests、python -m unittest tests.xxx、直接运行本文件三种方式
    from ._fixtures import SAMPLE_FILE, requires_sample_file
except ImportError:
    from _fixtures import SAMPLE_FILE, requires_sample_file

from openpyxl import load_workbook

from qpcr_tool.analysis import AnalysisError, analyze
from qpcr_tool.exporter import ExportError, build_wide_table, export_excel
from qpcr_tool.grouping import auto_group, strip_trailing_index
from qpcr_tool.reader import read_plate
from qpcr_tool.replicates import group_by_original_sample, split_biological_replicates

EXPECTED_TARGETS = ["GAPDH", "IL1b", "IL6", "TNFa"]
EXPECTED_SAMPLES = ["CT", "PBS BALF", "LPS BALF", "LPS BALF CIT013"]
REFERENCE = ["GAPDH"]
CONTROL = "CT"


def load_plate():
    """每次都重新读盘，避免某个用例改了 included 影响到别的用例。"""
    return read_plate(str(SAMPLE_FILE))


def run_pipeline():
    """读表 + 自动分组 + 计算，返回 (plate, groups, result)。"""
    plate = load_plate()
    groups = auto_group(plate.samples)
    result = analyze(plate.wells, groups, REFERENCE, CONTROL)
    return plate, groups, result


@requires_sample_file
class TestReader(unittest.TestCase):
    """任务 1：样本文件能被正确解析。"""

    def test_sample_file_exists(self):
        self.assertTrue(SAMPLE_FILE.is_file(), f"样本文件不存在：{SAMPLE_FILE}")

    def test_targets_samples_and_well_counts(self):
        plate = load_plate()
        self.assertEqual(plate.targets, EXPECTED_TARGETS)
        self.assertEqual(plate.samples, EXPECTED_SAMPLES)
        self.assertEqual(len(plate.wells), 48)
        self.assertEqual(len(plate.invalid_wells), 0)
        self.assertEqual(sum(1 for w in plate.wells if w.usable), 48)

    def test_每个基因每个样本都是三复孔(self):
        plate = load_plate()
        for target in EXPECTED_TARGETS:
            for sample in EXPECTED_SAMPLES:
                n = sum(1 for w in plate.wells if w.target == target and w.sample == sample)
                self.assertEqual(n, 3, f"{target} / {sample} 的孔数不是 3")


class TestGrouping(unittest.TestCase):
    """任务 2：自动分组。只有读盘的两条挂了跳过，纯字符串逻辑始终执行。"""

    @requires_sample_file
    def test_default_one_sample_per_group(self):
        plate = load_plate()
        groups = auto_group(plate.samples)
        self.assertEqual(len(groups), 4)
        self.assertEqual([g.name for g in groups], EXPECTED_SAMPLES)
        for group in groups:
            self.assertEqual(len(group.samples), 1)

    def test_merge_trailing_numbers(self):
        groups = auto_group(
            ["LPS-1", "LPS-2", "LPS-3", "PBS-1", "PBS-2"], merge_trailing_numbers=True
        )
        self.assertEqual([g.name for g in groups], ["LPS", "PBS"])
        self.assertEqual(groups[0].samples, ["LPS-1", "LPS-2", "LPS-3"])
        self.assertEqual(groups[1].samples, ["PBS-1", "PBS-2"])

    def test_strip_trailing_index_keeps_base_name_intact(self):
        """回归用例：rep/r/n/s 标记不能把基名最后一个字母一起吃掉。"""
        for raw, expected in [
            ("LPS-1", "LPS"), ("PBS-2", "PBS"), ("LPS_2", "LPS"), ("LPS 3", "LPS"),
            ("LPS#4", "LPS"), ("LPS(5)", "LPS"), ("LPS-rep2", "LPS"),
            ("LPS_S1", "LPS"), ("Ctrl-rep10", "Ctrl"),
            ("LPS", "LPS"), ("PBS BALF", "PBS BALF"), ("CT", "CT"),
        ]:
            self.assertEqual(strip_trailing_index(raw), expected, f"输入 {raw!r}")

    def test_default_mode_does_not_touch_sample_names(self):
        """默认模式不剥编号，CIT013 这种名字必须原样保留。"""
        groups = auto_group(["LPS BALF CIT013", "LPS BALF CIT013", "CT"])
        self.assertEqual([g.name for g in groups], ["LPS BALF CIT013", "CT"])

    def test_剥离后并不能合并的名字保持原样(self):
        """回归：误开合并选项时，LPS BALF CIT013 不能被静默剥成 LPS BALF CIT。"""
        groups = auto_group(
            ["LPS BALF CIT013", "CT", "PBS BALF"], merge_trailing_numbers=True
        )
        self.assertEqual(
            [g.name for g in groups], ["LPS BALF CIT013", "CT", "PBS BALF"]
        )
        self.assertEqual([g.samples for g in groups], [["LPS BALF CIT013"], ["CT"], ["PBS BALF"]])

    def test_真的能合并时照常合并(self):
        groups = auto_group(["LPS-1", "LPS-2", "LPS-3", "CT"], merge_trailing_numbers=True)
        self.assertEqual([g.name for g in groups], ["LPS", "CT"])
        self.assertEqual(groups[0].samples, ["LPS-1", "LPS-2", "LPS-3"])
        self.assertEqual(groups[1].samples, ["CT"])

    def test_同一批里既有能合并的也有不能合并的(self):
        groups = auto_group(
            ["LPS BALF CIT013", "PBS-1", "PBS-2", "CT013"], merge_trailing_numbers=True
        )
        self.assertEqual([g.name for g in groups], ["LPS BALF CIT013", "PBS", "CT013"])

    @requires_sample_file
    def test_合并模式下真实样本名不受影响(self):
        plate = load_plate()
        groups = auto_group(plate.samples, merge_trailing_numbers=True)
        self.assertEqual([g.name for g in groups], EXPECTED_SAMPLES)


@requires_sample_file
class TestAnalysis(unittest.TestCase):
    """任务 3、4、7：计算结果、手算校验、缺内参边界。"""

    def test_targets_exclude_reference(self):
        _, _, result = run_pipeline()
        self.assertEqual(result.targets, ["IL1b", "IL6", "TNFa"])
        self.assertEqual(result.groups, EXPECTED_SAMPLES)
        self.assertEqual(result.control_group, CONTROL)
        self.assertEqual(len(result.well_results), 3 * 4 * 3)

    def test_every_cell_has_three_replicates(self):
        _, _, result = run_pipeline()
        for target in result.targets:
            for group in result.groups:
                self.assertEqual(
                    len(result.values(target, group)), 3, f"{target} / {group}"
                )

    def test_control_group_rq_mean_close_to_one(self):
        """对照组 RQ 均值≈1，但数学上不会精确等于 1。

        ΔΔCt 在对照组内的算术平均恒为 0，所以 RQ 的**几何**平均精确等于 1；
        而 2^-x 是凸函数，由 Jensen 不等式，**算术**平均必定 ≥ 1 且通常略大于 1。
        因此算术平均只能用容差断言，几何平均才能断到高精度。
        """
        _, _, result = run_pipeline()
        for target in result.targets:
            values = result.values(target, CONTROL)
            arithmetic = sum(values) / len(values)
            geometric = math.exp(sum(math.log(v) for v in values) / len(values))
            self.assertAlmostEqual(geometric, 1.0, places=9, msg=f"{target} 几何平均")
            self.assertGreaterEqual(arithmetic, 1.0 - 1e-12, f"{target} 违反 Jensen 不等式")
            self.assertAlmostEqual(arithmetic, 1.0, delta=0.05, msg=f"{target} 算术平均")

    def test_manual_recalculation_il1b_lps(self):
        """独立手算校验：不调用 analysis.py 的任何函数生成期望值。

        直接把 IL1b / LPS BALF 三个孔的 Cq 硬编码进来，用最朴素的
        sum()/len() 重新走一遍 ref_mean → ΔCt → ΔΔCt → 2^-ΔΔCt。
        """
        # 下机表原始 Cq（GAPDH 为内参，CT 为对照组）
        gapdh_ct = [17.6980823376913, 18.0449978324619, 17.8214612438022]  # A01-A03
        il1b_ct = [16.35886281486, 16.5606297727241, 16.6662063969497]     # A04-A06
        gapdh_lps = [17.2131293783367, 17.2330959151691, 17.320584091062]  # C01-C03
        il1b_lps = [14.9657957365361, 15.0921062502669, 15.2262013737718]  # C04-C06

        ref_mean_ct = sum(gapdh_ct) / len(gapdh_ct)
        ref_mean_lps = sum(gapdh_lps) / len(gapdh_lps)
        baseline = sum(cq - ref_mean_ct for cq in il1b_ct) / len(il1b_ct)
        expected_rq = [2.0 ** (-((cq - ref_mean_lps) - baseline)) for cq in il1b_lps]

        _, _, result = run_pipeline()
        actual = [
            item for item in result.well_results
            if item.target == "IL1b" and item.sample == "LPS BALF"
        ]
        self.assertEqual([item.well for item in actual], ["C04", "C05", "C06"])
        self.assertAlmostEqual(actual[0].ref_mean, ref_mean_lps, delta=1e-9)

        for item, expected in zip(actual, expected_rq):
            self.assertAlmostEqual(item.rq, expected, delta=1e-9, msg=f"孔 {item.well}")
            self.assertAlmostEqual(
                item.dct, item.cq - ref_mean_lps, delta=1e-9, msg=f"孔 {item.well} ΔCt"
            )
            self.assertAlmostEqual(
                item.ddct, (item.cq - ref_mean_lps) - baseline, delta=1e-9,
                msg=f"孔 {item.well} ΔΔCt",
            )

    def test_missing_reference_raises(self):
        """把所有 GAPDH 孔剔除后没有内参可用，analyze 应当直接报错而不是算出垃圾。"""
        plate = load_plate()
        for well in plate.wells:
            if well.target == "GAPDH":
                well.included = False
        groups = auto_group(plate.samples)
        with self.assertRaises(AnalysisError) as ctx:
            analyze(plate.wells, groups, REFERENCE, CONTROL)
        self.assertIn("无法归一化", str(ctx.exception))


@requires_sample_file
class TestWideTable(unittest.TestCase):
    """任务 5：宽表结构。"""

    def setUp(self):
        _, _, self.result = run_pipeline()
        self.table = build_wide_table(self.result)

    def test_shape(self):
        self.assertEqual(len(self.table.columns), 3 * 4)
        self.assertEqual(len(self.table.rows), 3)
        for row in self.table.rows:
            self.assertEqual(len(row), 12)
            self.assertTrue(all(isinstance(v, float) for v in row))

    def test_column_order_is_gene_outer_group_inner(self):
        self.assertEqual(
            self.table.columns[:4],
            [("IL1b", s) for s in EXPECTED_SAMPLES],
        )
        self.assertEqual(self.table.gene_header[:4], ["IL1b"] * 4)
        self.assertEqual(self.table.group_header[:4], EXPECTED_SAMPLES)

    def test_merge_spans(self):
        spans = self.table.merge_spans()
        self.assertEqual(len(spans), 3)
        self.assertEqual([gene for _, _, gene in spans], ["IL1b", "IL6", "TNFa"])
        for start, end, _ in spans:
            self.assertEqual(end - start + 1, 4)
        self.assertEqual(spans[0][:2], (0, 3))
        self.assertEqual(spans[2][:2], (8, 11))

    def test_to_tsv(self):
        lines = self.table.to_tsv().split("\n")
        self.assertEqual(len(lines), 2 + 3)
        for line in lines:
            self.assertEqual(len(line.split("\t")), 12)
        # 基因名只出现在每个块的第一列
        self.assertEqual(
            lines[0].split("\t"), ["IL1b", "", "", "", "IL6", "", "", "", "TNFa", "", "", ""]
        )
        self.assertEqual(lines[1].split("\t"), EXPECTED_SAMPLES * 3)
        self.assertEqual(len(self.table.to_tsv(include_header=False).split("\n")), 3)

    def test_short_column_padded_with_none(self):
        """某组孔数不足时，该列后面的行必须补 None 而不是错位。"""
        self.result.stats[("IL1b", "CT")].values.pop()
        table = build_wide_table(self.result)
        self.assertEqual(len(table.rows), 3)
        self.assertIsNone(table.rows[2][0])
        self.assertIsNotNone(table.rows[2][1])
        # 缺值在 TSV 里是空字段，不能让后面的数字往前顶造成错列
        fields = table.to_tsv().split("\n")[-1].split("\t")
        self.assertEqual(len(fields), 12)
        self.assertEqual(fields[0], "")
        self.assertNotEqual(fields[1], "")


@requires_sample_file
class TestExportExcel(unittest.TestCase):
    """任务 6：Excel 导出后能被读回并对得上。"""

    @classmethod
    def setUpClass(cls):
        cls.plate, _, cls.result = run_pipeline()
        cls.table = build_wide_table(cls.result)
        cls.tmpdir = tempfile.TemporaryDirectory()
        cls.path = os.path.join(cls.tmpdir.name, "导出测试.xlsx")
        export_excel(cls.result, cls.plate, cls.path)
        cls.workbook = load_workbook(cls.path)

    @classmethod
    def tearDownClass(cls):
        cls.workbook.close()
        cls.tmpdir.cleanup()

    def test_sheets_exist(self):
        self.assertEqual(self.workbook.sheetnames, ["Prism", "明细QC"])

    def test_prism_header(self):
        sheet = self.workbook["Prism"]
        self.assertEqual(sheet["A1"].value, "IL1b")
        self.assertEqual(sheet["E1"].value, "IL6")
        self.assertEqual(sheet["I1"].value, "TNFa")
        self.assertIsNone(sheet["B1"].value)  # 被合并进 A1
        self.assertEqual(sheet["A2"].value, "CT")
        self.assertEqual(sheet["B2"].value, "PBS BALF")
        self.assertEqual(sheet["C2"].value, "LPS BALF")
        self.assertEqual(sheet["D2"].value, "LPS BALF CIT013")
        self.assertEqual(sheet["E2"].value, "CT")
        self.assertEqual(len(sheet.merged_cells.ranges), 3)
        self.assertEqual(sheet.freeze_panes, "A3")

    def test_prism_values_are_raw_floats(self):
        """写进去的必须是 float 原值，显示精度只靠 number_format 控制。

        openpyxl 落盘时按 %.16g 序列化（Excel 自身也只保 15 位有效数字），
        所以读回来可能和内存值差最后一位，这里用相对容差而不是精确相等。
        """
        sheet = self.workbook["Prism"]
        self.assertEqual(sheet.max_row, 5)
        self.assertEqual(sheet.max_column, 12)
        beyond_four_decimals = 0
        for row_offset, row in enumerate(self.table.rows, start=3):
            for col_offset, expected in enumerate(row, start=1):
                cell = sheet.cell(row=row_offset, column=col_offset)
                self.assertIsInstance(cell.value, float)
                self.assertEqual(cell.number_format, "0.0000")
                self.assertAlmostEqual(cell.value, expected, delta=abs(expected) * 1e-14)
                if cell.value != round(expected, 4):
                    beyond_four_decimals += 1
        # 36 个值都保留了四位小数以外的精度，说明没有被提前 round 成显示值
        self.assertEqual(beyond_four_decimals, 36)

    def test_detail_sheet_covers_every_well(self):
        """所有 48 个孔都要在明细里出现，包括内参孔，不能静默丢数据。"""
        sheet = self.workbook["明细QC"]
        listed = {
            (row[0], row[1], row[2])
            for row in sheet.iter_rows(min_col=1, max_col=3, values_only=True)
        }
        for well in self.plate.wells:
            self.assertIn(
                (well.well, well.target, well.sample), listed, f"{well.well} 未出现在明细里"
            )

    def test_detail_sheet_sections_and_reference_reason(self):
        sheet = self.workbook["明细QC"]
        rows = list(sheet.iter_rows(values_only=True))
        column_a = [row[0] for row in rows]
        for title in ("逐孔明细", "分组汇总", "分析参数"):
            self.assertIn(title, column_a)
        self.assertTrue(
            any(isinstance(v, str) and v.startswith("复孔一致性") for v in column_a)
        )

        # 只看「逐孔明细」表头后面的 48 行，别把「分析参数」里的 GAPDH 也数进来
        header_idx = column_a.index("逐孔明细") + 1
        self.assertEqual(rows[header_idx][0], "孔位")
        detail = rows[header_idx + 1: header_idx + 1 + len(self.plate.wells)]
        self.assertEqual(len(detail), 48)
        self.assertEqual(len([r for r in detail if r[5] == "是"]), 36)

        gapdh_rows = [r for r in detail if r[1] == "GAPDH"]
        self.assertEqual(len(gapdh_rows), 12)
        for row in gapdh_rows:
            self.assertTrue(str(row[5]).startswith("否"))
            self.assertIn("内参基因", str(row[5]))

    def test_detail_sheet_parameters(self):
        sheet = self.workbook["明细QC"]
        params = {
            row[0]: row[1]
            for row in sheet.iter_rows(min_col=1, max_col=2, values_only=True)
            if isinstance(row[0], str)
        }
        self.assertEqual(params["源文件名"], os.path.basename(self.plate.file_path))
        self.assertEqual(params["数据所在 sheet"], self.plate.sheet_name)
        self.assertEqual(params["内参基因"], "GAPDH")
        self.assertEqual(params["对照组"], CONTROL)
        self.assertIn("ΔCt_i", str(params["计算公式 ①"]))
        self.assertIn("ΔΔCt_i", str(params["计算公式 ②"]))
        self.assertIn("2^(−ΔΔCt_i)", str(params["计算公式 ③"]))
        from qpcr_tool import __version__

        self.assertEqual(params["软件版本"], __version__)

    def test_group_summary_matches_stats(self):
        sheet = self.workbook["明细QC"]
        rows = list(sheet.iter_rows(values_only=True))
        header_idx = next(i for i, r in enumerate(rows) if r[0] == "基因" and r[3] == "均值")
        seen = 0
        for row in rows[header_idx + 1: header_idx + 1 + 12]:
            stat = self.result.stat(row[0], row[1])
            self.assertIsNotNone(stat, f"{row[0]} / {row[1]}")
            self.assertEqual(row[2], stat.n)
            self.assertAlmostEqual(row[3], stat.mean, delta=1e-12)
            self.assertAlmostEqual(row[4], stat.sd, delta=1e-12)
            seen += 1
        self.assertEqual(seen, 12)

    def test_sections_separated_by_blank_row(self):
        sheet = self.workbook["明细QC"]
        rows = list(sheet.iter_rows(values_only=True))
        # 小标题行的特征：只有 A 列有内容，右边全空
        titles = [
            (index, row[0]) for index, row in enumerate(rows)
            if isinstance(row[0], str) and all(value is None for value in row[1:])
        ]
        self.assertEqual(len(titles), 4)
        self.assertEqual(titles[0][1], "逐孔明细")
        self.assertEqual(titles[1][1], "分组汇总")
        self.assertTrue(titles[2][1].startswith("复孔一致性"))
        self.assertEqual(titles[3][1], "分析参数")
        for index, name in titles[1:]:
            self.assertTrue(
                all(value is None for value in rows[index - 1]),
                f"小标题「{name}」前面没有空行",
            )

    def test_export_error_on_unwritable_path(self):
        """目标路径写不进去（被 Excel 占用、或就是个目录）时要给友好报错。"""
        with self.assertRaises(ExportError) as ctx:
            export_excel(self.result, self.plate, self.tmpdir.name)
        self.assertTrue(str(ctx.exception))


@requires_sample_file
class TestExportWithWarnings(unittest.TestCase):
    """对照组缺某个基因时：该基因整体跳过、产生警告，导出要如实反映。"""

    @classmethod
    def setUpClass(cls):
        cls.plate = load_plate()
        for well in cls.plate.wells:
            if well.target == "IL6" and well.sample == CONTROL:
                well.included = False
        groups = auto_group(cls.plate.samples)
        cls.result = analyze(cls.plate.wells, groups, REFERENCE, CONTROL)
        cls.tmpdir = tempfile.TemporaryDirectory()
        cls.path = os.path.join(cls.tmpdir.name, "warn.xlsx")
        export_excel(cls.result, cls.plate, cls.path)
        cls.workbook = load_workbook(cls.path)

    @classmethod
    def tearDownClass(cls):
        cls.workbook.close()
        cls.tmpdir.cleanup()

    def test_gene_dropped_and_warned(self):
        self.assertEqual(self.result.targets, ["IL1b", "TNFa"])
        self.assertTrue(any("IL6" in w for w in self.result.warnings))

    def test_prism_shrinks_to_two_genes(self):
        sheet = self.workbook["Prism"]
        self.assertEqual(sheet.max_column, 2 * 4)
        self.assertEqual(len(sheet.merged_cells.ranges), 2)

    def test_warning_section_written(self):
        sheet = self.workbook["明细QC"]
        column_a = [row[0] for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True)]
        self.assertIn("分析警告", column_a)
        text = "\n".join(
            str(row[1]) for row in sheet.iter_rows(min_col=1, max_col=2, values_only=True)
        )
        self.assertIn("IL6", text)

    def test_excluded_il6_wells_have_reasons(self):
        """IL6 的 12 个孔全部退出结果，理由要能区分手动剔除和整基因跳过。"""
        sheet = self.workbook["明细QC"]
        rows = list(sheet.iter_rows(values_only=True))
        il6 = [r for r in rows if r[1] == "IL6" and isinstance(r[5], str)]
        self.assertEqual(len(il6), 12)
        self.assertEqual(sum(1 for r in il6 if "手动剔除" in r[5]), 3)
        self.assertEqual(sum(1 for r in il6 if "该基因在对照组无数据" in r[5]), 9)


@requires_sample_file
class TestExportWithSplitReport(unittest.TestCase):
    """任务 5：拆分信息要写进明细表，读回来能对上。"""

    @classmethod
    def setUpClass(cls):
        cls.plate = load_plate()
        cls.report = split_biological_replicates(cls.plate)
        cls.result = analyze(
            cls.plate.wells, group_by_original_sample(cls.plate), REFERENCE, CONTROL
        )
        cls.tmpdir = tempfile.TemporaryDirectory()
        cls.path = os.path.join(cls.tmpdir.name, "拆分导出.xlsx")
        export_excel(cls.result, cls.plate, cls.path, cls.report)
        cls.workbook = load_workbook(cls.path)
        cls.rows = list(cls.workbook["明细QC"].iter_rows(values_only=True))

    @classmethod
    def tearDownClass(cls):
        cls.workbook.close()
        cls.tmpdir.cleanup()

    def detail_rows(self) -> list[tuple]:
        column_a = [row[0] for row in self.rows]
        header_idx = column_a.index("逐孔明细") + 1
        return self.rows[header_idx: header_idx + 1 + len(self.plate.wells)]

    def parameters(self) -> dict[str, str]:
        return {
            row[0]: row[1] for row in self.rows if isinstance(row[0], str)
        }

    def test_逐孔明细多了原样本名和重复编号两列(self):
        detail = self.detail_rows()
        self.assertEqual(detail[0][10], "原样本名")
        self.assertEqual(detail[0][11], "重复编号")
        by_well = {row[0]: row for row in detail[1:]}
        self.assertEqual(by_well["A04"][2], "CT-1")
        self.assertEqual(by_well["A04"][10], "CT")
        self.assertEqual(by_well["A04"][11], "1")
        self.assertEqual(by_well["E06"][10], "CT")
        self.assertEqual(by_well["E06"][11], "3")
        # 内参孔虽然不进结果，也要带上归属信息
        self.assertEqual(by_well["A03"][10], "CT")
        self.assertEqual(by_well["A03"][11], "3")

    def test_原有列的位置没有被挤动(self):
        detail = self.detail_rows()
        self.assertEqual(
            list(detail[0][:10]),
            ["孔位", "基因", "样本", "分组", "Cq", "参与计算", "内参均值", "ΔCt", "ΔΔCt", "2^-ΔΔCt"],
        )
        self.assertEqual(len([r for r in detail[1:] if r[5] == "是"]), 36)

    def test_分析参数记录了拆分口径(self):
        params = self.parameters()
        self.assertEqual(params["复孔处理模式"], "生物学重复配对（同名样本已拆分）")
        self.assertIn("自动识别", params["孔位排列方向"])
        self.assertIn("横向排列 16 个区块", params["孔位排列方向"])
        self.assertEqual(
            params["已拆分的样本"], "CT、PBS BALF、LPS BALF、LPS BALF CIT013"
        )

    def test_复孔一致性恢复成十六行三复孔(self):
        column_a = [row[0] for row in self.rows]
        title_idx = next(
            i for i, v in enumerate(column_a) if isinstance(v, str) and v.startswith("复孔一致性")
        )
        header = self.rows[title_idx + 1]
        self.assertEqual(list(header[:6]), ["基因", "样本", "n", "Cq 均值", "Cq 标准差", "是否超阈值"])
        body = self.rows[title_idx + 2: title_idx + 2 + 16]
        self.assertEqual(len(body), 16)
        for row in body:
            self.assertEqual(row[2], 3, f"{row[0]} / {row[1]} 不是 3 复孔")
            self.assertIsNotNone(row[4], f"{row[0]} / {row[1]} 的标准差是空的")
        self.assertEqual({row[1] for row in body}, set(EXPECTED_SAMPLES))

    def test_拆分提示会被写进参数段(self):
        """孔数不一致这类提示要留在报告里，不能只在界面上一闪而过。"""
        plate = load_plate()
        plate.wells = [w for w in plate.wells if w.well != "A03"]
        report = split_biological_replicates(plate)
        result = analyze(
            plate.wells, group_by_original_sample(plate), REFERENCE, CONTROL
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "提示.xlsx")
            export_excel(result, plate, path, report)
            workbook = load_workbook(path)
            try:
                params = {
                    row[0]: row[1]
                    for row in workbook["明细QC"].iter_rows(min_col=1, max_col=2, values_only=True)
                    if isinstance(row[0], str)
                }
            finally:
                workbook.close()
        self.assertIn("各基因孔数不一致", str(params["拆分提示 1"]))
        self.assertEqual(params["已拆分的样本"], "PBS BALF、LPS BALF、LPS BALF CIT013")

    def test_不传拆分报告时按技术复孔如实记录(self):
        plate = load_plate()
        result = analyze(plate.wells, auto_group(plate.samples), REFERENCE, CONTROL)
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "未拆分.xlsx")
            export_excel(result, plate, path)
            workbook = load_workbook(path)
            try:
                rows = list(workbook["明细QC"].iter_rows(values_only=True))
            finally:
                workbook.close()
        params = {row[0]: row[1] for row in rows if isinstance(row[0], str)}
        self.assertEqual(params["复孔处理模式"], "技术复孔取平均")
        self.assertIn("未启用", params["孔位排列方向"])
        self.assertEqual(params["已拆分的样本"], "（无）")
        header_idx = [row[0] for row in rows].index("逐孔明细") + 1
        detail = rows[header_idx + 1: header_idx + 1 + 48]
        self.assertEqual({row[11] for row in detail}, {"-"})
        self.assertEqual({row[10] for row in detail}, set(EXPECTED_SAMPLES))


@requires_sample_file
class TestReplicateFlagStyling(unittest.TestCase):
    """复孔离散度超阈值的行必须标黄，否则 QC 提示在 Excel 里等于没有。"""

    def test_flagged_row_is_highlighted(self):
        plate = load_plate()
        for well in plate.wells:
            if well.well == "A06":  # IL1b / CT，人为把 Cq 拉偏制造高 SD
                well.cq = 20.0
        result = analyze(plate.wells, auto_group(plate.samples), REFERENCE, CONTROL)
        flagged = [(q.target, q.sample) for q in result.qc if q.flagged]
        self.assertEqual(flagged, [("IL1b", CONTROL)])

        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "flag.xlsx")
            export_excel(result, plate, path)
            workbook = load_workbook(path)
            try:
                sheet = workbook["明细QC"]
                marked = [
                    row for row in sheet.iter_rows(min_col=1, max_col=6)
                    if row[5].value == "⚠ 偏高"
                ]
                self.assertEqual(len(marked), 1)
                self.assertEqual(marked[0][0].value, "IL1b")
                for cell in marked[0]:
                    self.assertEqual(cell.fill.patternType, "solid")
                    self.assertEqual(cell.fill.fgColor.rgb, "FFFFEB9C")
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main(verbosity=2)
