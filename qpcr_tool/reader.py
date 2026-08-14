"""读取 qPCR 下机 Excel，自动定位表头并解析出孔记录。

不写死列位置，因此 Bio-Rad CFX、ABI QuantStudio / 7500 等导出的表格都能直接解析。
"""

from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass, field
from typing import Any, BinaryIO

from openpyxl import load_workbook

# Bio-Rad CFX 偶发导出小写 ZIP 条目名（如 [content_types].xml / sharedstrings.xml）。
# Windows 上 Excel 不区分大小写所以能开，但 openpyxl 严格按 OOXML 规范查找会失败。
# 这里把已知关键路径归一化成规范大小写后再交给 openpyxl。
_OOXML_CANONICAL = {
    "[content_types].xml": "[Content_Types].xml",
    "xl/sharedstrings.xml": "xl/sharedStrings.xml",
}

HEADER_SEARCH_ROWS = 40

# 表头别名。全部转小写去空格后匹配。
WELL_ALIASES = ("well", "wellposition", "position", "孔", "孔位")
TARGET_ALIASES = (
    "target", "targetname", "detector", "detectorname",
    "gene", "genename", "assay", "检测目标", "基因",
)
SAMPLE_ALIASES = (
    "sample", "samplename", "samplelabel", "sampleid",
    "样本", "样品", "样本名称",
)
CQ_ALIASES = ("cq", "ct", "cp", "cqvalue", "ctvalue", "crossingpoint")
CONTENT_ALIASES = ("content", "task", "tasktype", "type")

# 命中这些词的列不能当作 Cq 主列（Cq Mean / Cq Std. Dev / SQ ... ）
CQ_EXCLUDE = ("mean", "std", "sd", "sq", "quantity", "conf", "delta", "threshold")

# 视为「未检出 / 无效」的 Cq 文本
INVALID_CQ_TEXT = (
    "", "n/a", "na", "nan", "undetermined", "undet", "-", "--", "none", "null",
)


def _norm(value: Any) -> str:
    """表头归一化：转小写、去掉空白与常见标点。"""
    if value is None:
        return ""
    return re.sub(r"[\s_.\-()（）\[\]]", "", str(value)).strip().lower()


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@dataclass
class WellRecord:
    """一个孔的原始记录。"""

    row_index: int
    well: str
    target: str
    sample: str
    content: str
    cq: float | None
    cq_text: str
    included: bool = True
    original_sample: str = ""    # 拆分前的样本名，拆分后用于回溯和分组
    replicate_index: int = 0     # 该孔在其样本内的生物学重复序号，1-based；0 表示尚未编号

    @property
    def valid(self) -> bool:
        return self.cq is not None

    @property
    def usable(self) -> bool:
        """真正参与计算：Cq 有效且用户没有剔除。"""
        return self.valid and self.included


@dataclass
class PlateData:
    """一份下机文件的解析结果。"""

    file_path: str
    sheet_name: str
    wells: list[WellRecord] = field(default_factory=list)
    run_info: list[tuple[str, str]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def targets(self) -> list[str]:
        """按首次出现顺序返回所有目标基因。"""
        return _unique([w.target for w in self.wells])

    @property
    def samples(self) -> list[str]:
        return _unique([w.sample for w in self.wells])

    @property
    def invalid_wells(self) -> list[WellRecord]:
        return [w for w in self.wells if not w.valid]


def _unique(items: list[str]) -> list[str]:
    seen: dict[str, None] = {}
    for item in items:
        if item:
            seen.setdefault(item, None)
    return list(seen)


class ReaderError(Exception):
    """文件无法解析时抛出，消息直接展示给用户。"""


def _match_column(header_cells: list[str], aliases: tuple[str, ...],
                  exclude: tuple[str, ...] = ()) -> int | None:
    """先找完全相等的列，再退回包含匹配。返回 0-based 列下标。"""
    for idx, cell in enumerate(header_cells):
        if cell in aliases:
            return idx
    for idx, cell in enumerate(header_cells):
        if not cell or any(bad in cell for bad in exclude):
            continue
        if any(cell.startswith(alias) or alias in cell for alias in aliases):
            return idx
    return None


def _find_header_row(rows: list[list[Any]]) -> tuple[int, dict[str, int]] | None:
    """在前若干行里找同时具备 Target/Sample/Cq 的表头行。"""
    for row_idx, row in enumerate(rows[:HEADER_SEARCH_ROWS]):
        cells = [_norm(c) for c in row]
        if not any(cells):
            continue
        target_col = _match_column(cells, TARGET_ALIASES)
        sample_col = _match_column(cells, SAMPLE_ALIASES)
        cq_col = _match_column(cells, CQ_ALIASES, exclude=CQ_EXCLUDE)
        if target_col is None or sample_col is None or cq_col is None:
            continue
        columns = {"target": target_col, "sample": sample_col, "cq": cq_col}
        well_col = _match_column(cells, WELL_ALIASES)
        if well_col is not None:
            columns["well"] = well_col
        content_col = _match_column(cells, CONTENT_ALIASES)
        if content_col is not None:
            columns["content"] = content_col
        return row_idx, columns
    return None


def _parse_cq(raw: Any) -> tuple[float | None, str]:
    """返回 (数值或 None, 原始文本)。0 与负值按未检出处理。"""
    text = _text(raw)
    if text.lower() in INVALID_CQ_TEXT:
        return None, text
    try:
        value = float(str(raw).strip())
    except (TypeError, ValueError):
        return None, text
    if value <= 0:
        return None, text
    return value, text


def _read_run_info(workbook) -> list[tuple[str, str]]:
    for name in workbook.sheetnames:
        if "run" in name.lower() and "info" in name.lower():
            info: list[tuple[str, str]] = []
            for row in workbook[name].iter_rows(values_only=True):
                if not row:
                    continue
                key = _text(row[0])
                value = _text(row[1]) if len(row) > 1 else ""
                if key:
                    info.append((key, value))
            return info
    return []


def _needs_ooxml_case_fix(file_path: str) -> bool:
    """ZIP 里是否有需要改回规范大小写的 OOXML 条目。"""
    try:
        with zipfile.ZipFile(file_path) as archive:
            names = {name.replace("\\", "/") for name in archive.namelist()}
    except zipfile.BadZipFile:
        return False
    for wrong, right in _OOXML_CANONICAL.items():
        if wrong in names and right not in names:
            return True
    # 已有规范名就不改；但若只有错误大小写的变体（路径大小写混乱）也要修
    lower_map = {name.lower(): name for name in names}
    for wrong, right in _OOXML_CANONICAL.items():
        if right.lower() in lower_map and lower_map[right.lower()] != right:
            return True
    return False


def _normalize_ooxml_casing(file_path: str) -> BinaryIO:
    """把 xlsx 包里的关键条目名改成 OOXML 规范大小写，返回可被 openpyxl 读取的内存流。"""
    buffer = io.BytesIO()
    with zipfile.ZipFile(file_path, "r") as src, zipfile.ZipFile(
        buffer, "w", compression=zipfile.ZIP_DEFLATED
    ) as dst:
        used: set[str] = set()
        for info in src.infolist():
            original = info.filename.replace("\\", "/")
            canonical = _OOXML_CANONICAL.get(original.lower(), original)
            # 同一规范名只写一次，避免大小写冲突产生重复条目
            if canonical.lower() in used:
                continue
            used.add(canonical.lower())
            dst.writestr(canonical, src.read(info.filename))
    buffer.seek(0)
    return buffer


def _load_workbook(file_path: str):
    """打开工作簿；必要时先修正 CFX 导出的小写 ZIP 条目名。"""
    if _needs_ooxml_case_fix(file_path):
        stream = _normalize_ooxml_casing(file_path)
        workbook = load_workbook(stream, data_only=True, read_only=True)
        # read_only 模式会一直握着底层流，必须把流挂在 workbook 上防止被回收
        workbook._qpcr_source_stream = stream  # type: ignore[attr-defined]
        return workbook
    return load_workbook(file_path, data_only=True, read_only=True)


def read_plate(file_path: str) -> PlateData:
    """解析一个 Excel 下机文件。找不到可用表头时抛 ReaderError。"""
    try:
        workbook = _load_workbook(file_path)
    except Exception as exc:  # openpyxl 的异常类型很杂，统一转成可读消息
        raise ReaderError(f"无法打开文件：{exc}") from exc

    try:
        parsed: tuple[str, int, dict[str, int], list[list[Any]]] | None = None
        for sheet_name in workbook.sheetnames:
            rows = [list(r) for r in workbook[sheet_name].iter_rows(values_only=True)]
            found = _find_header_row(rows)
            if found is not None:
                header_idx, columns = found
                parsed = (sheet_name, header_idx, columns, rows)
                break

        if parsed is None:
            raise ReaderError(
                "没有在这个文件里找到 qPCR 数据表。\n"
                "需要有一行表头同时包含「Target」「Sample」和「Cq / Ct」三列。"
            )

        sheet_name, header_idx, columns, rows = parsed
        run_info = _read_run_info(workbook)
    finally:
        workbook.close()

    plate = PlateData(file_path=file_path, sheet_name=sheet_name, run_info=run_info)

    def cell(row: list[Any], key: str) -> Any:
        idx = columns.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for offset, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        target = _text(cell(row, "target"))
        sample = _text(cell(row, "sample"))
        if not target and not sample:
            continue  # 空孔
        cq_value, cq_text = _parse_cq(cell(row, "cq"))
        well_name = _text(cell(row, "well")) or f"R{offset}"
        plate.wells.append(
            WellRecord(
                row_index=offset,
                well=well_name,
                target=target or "(未命名基因)",
                sample=sample or "(未命名样本)",
                content=_text(cell(row, "content")),
                cq=cq_value,
                cq_text=cq_text,
                included=cq_value is not None,
                original_sample=sample or "(未命名样本)",
            )
        )

    if not plate.wells:
        raise ReaderError("表头找到了，但下面没有任何带 Target / Sample 的数据行。")

    missing = [w for w in plate.wells if not w.valid]
    if missing:
        plate.warnings.append(
            f"有 {len(missing)} 个孔的 Cq 无效或未检出，已默认不参与计算。"
        )

    return plate
