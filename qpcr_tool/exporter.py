"""把分析结果导出成 Prism 宽表和 Excel 报告。

两个产物：
- ``WideTable``：内存里的宽表结构，列 = 基因 × 分组、行 = 各复孔的 2^-ΔΔCt，
  界面预览、剪贴板复制、Excel 导出共用同一份数据。
- ``export_excel``：写出双 sheet 的 xlsx。``Prism`` sheet 只放表头和数值，
  方便整块框选粘进 Prism；``明细QC`` sheet 记录逐孔链路、分组汇总、
  复孔一致性和分析参数，保证任何一个孔都能被追溯，不静默丢数据。
"""

from __future__ import annotations

import os
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import __version__
from .analysis import CQ_SD_WARN, AnalysisResult
from .reader import PlateData, WellRecord
from .replicates import ORIENTATION_LABELS, SplitReport

# to_tsv 输出的小数位。Excel 里数值是原值、只靠 number_format 控制显示，
# 但纯文本没有这一层，所以这里固定一个既够用又不刺眼的精度。
TSV_DECIMALS = 6

# 数值格式：Cq / ΔCt / ΔΔCt / RQ 统一四位小数
NUM_FORMAT = "0.0000"
INT_FORMAT = "0"

_THIN = Side(style="thin", color="FFB0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_BOLD = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=12)
_HEADER_FILL = PatternFill("solid", fgColor="FFDCE6F1")
_WARN_FILL = PatternFill("solid", fgColor="FFFFEB9C")


class ExportError(Exception):
    """导出失败时抛出，消息直接展示给用户。"""


@dataclass
class WideTable:
    """给 Prism 用的宽表：列 = 基因 × 分组，行 = 各复孔的 2^-ΔΔCt。"""

    columns: list[tuple[str, str]] = field(default_factory=list)
    rows: list[list[float | None]] = field(default_factory=list)

    @property
    def gene_header(self) -> list[str]:
        """每列对应的基因名。"""
        return [gene for gene, _ in self.columns]

    @property
    def group_header(self) -> list[str]:
        """每列对应的组名。"""
        return [group for _, group in self.columns]

    def merge_spans(self) -> list[tuple[int, int, str]]:
        """基因表头的合并区间，(起始列下标, 结束列下标, 基因名)，0-based 闭区间。

        按相邻同名基因合并，因此列顺序变化时不会把不连续的同名块错误地并到一起。
        """
        spans: list[tuple[int, int, str]] = []
        start = 0
        genes = self.gene_header
        for index in range(1, len(genes) + 1):
            if index == len(genes) or genes[index] != genes[start]:
                spans.append((start, index - 1, genes[start]))
                start = index
        return spans

    def to_tsv(self, include_header: bool = True) -> str:
        """制表符分隔的纯文本，行用 "\\n" 连接。

        include_header 为真时前两行分别是基因名行和组名行；基因名只写在每个
        基因块的第一列，块内其余列留空，和 Excel 里的跨列合并视觉一致。
        缺值输出空字符串。
        """
        lines: list[str] = []
        if include_header:
            gene_row = [""] * len(self.columns)
            for start, _end, gene in self.merge_spans():
                gene_row[start] = gene
            lines.append("\t".join(gene_row))
            lines.append("\t".join(self.group_header))
        for row in self.rows:
            lines.append(
                "\t".join(
                    "" if value is None else f"{value:.{TSV_DECIMALS}f}"
                    for value in row
                )
            )
        return "\n".join(lines)


def build_wide_table(result: AnalysisResult) -> WideTable:
    """把分析结果摊平成宽表。

    列顺序：外层遍历 ``result.targets``、内层遍历 ``result.groups``。
    行数取所有列里最长的那组孔数，孔数不足的列在后面补 None。
    """
    columns = [(target, group) for target in result.targets for group in result.groups]
    column_values = [result.values(target, group) for target, group in columns]
    row_count = max((len(values) for values in column_values), default=0)
    rows: list[list[float | None]] = [
        [values[i] if i < len(values) else None for values in column_values]
        for i in range(row_count)
    ]
    return WideTable(columns=columns, rows=rows)


def export_excel(
    result: AnalysisResult,
    plate: PlateData,
    path: str,
    split_report: SplitReport | None = None,
) -> None:
    """导出双 sheet 的 Excel 报告。目标文件被占用时抛 ExportError。

    ``split_report`` 传入时，明细表会额外记录每个孔属于哪只动物、本次用的复孔处理
    模式和孔位排列方向；不传表示按技术复孔取平均，报告里会如实写明。
    """
    workbook = Workbook()
    prism_sheet = workbook.active
    prism_sheet.title = "Prism"
    _write_prism_sheet(prism_sheet, build_wide_table(result))
    _write_detail_sheet(workbook.create_sheet("明细QC"), result, plate, split_report)

    try:
        workbook.save(path)
    except PermissionError as exc:
        raise ExportError(
            f"无法写入「{path}」。\n文件可能正被 Excel 打开，请关闭后重试。"
        ) from exc
    except OSError as exc:
        raise ExportError(f"保存「{path}」失败：{exc}") from exc
    finally:
        workbook.close()


# --------------------------------------------------------------------------
# Sheet 1：Prism
# --------------------------------------------------------------------------


def _write_prism_sheet(sheet: Worksheet, table: WideTable) -> None:
    """写纯数值区：第 1 行基因（跨列合并）、第 2 行分组、第 3 行起 RQ。"""
    # 先给整行铺样式再合并：合并之后非左上角的格子会变成只读的 MergedCell，
    # 但边框仍要逐格设置，否则合并区域画不全外框。
    for index in range(1, len(table.columns) + 1):
        cell = sheet.cell(row=1, column=index)
        cell.font = _BOLD
        cell.alignment = _CENTER
        cell.border = _BORDER

    for start, end, gene in table.merge_spans():
        sheet.cell(row=1, column=start + 1).value = gene
        if end > start:  # openpyxl 不接受单格合并
            sheet.merge_cells(
                start_row=1, start_column=start + 1, end_row=1, end_column=end + 1
            )

    for index, group in enumerate(table.group_header, start=1):
        cell = sheet.cell(row=2, column=index, value=group)
        cell.font = _BOLD
        cell.alignment = _CENTER
        cell.fill = _HEADER_FILL
        cell.border = _BORDER

    for row_offset, row in enumerate(table.rows, start=3):
        for col_offset, value in enumerate(row, start=1):
            # 写 float 原值，显示精度交给 number_format，避免提前 round 丢精度
            cell = sheet.cell(row=row_offset, column=col_offset, value=value)
            cell.number_format = NUM_FORMAT
            cell.alignment = _CENTER
            cell.border = _BORDER

    sheet.freeze_panes = "A3"
    _autofit(sheet, min_width=14, max_width=18)


# --------------------------------------------------------------------------
# Sheet 2：明细QC
# --------------------------------------------------------------------------


def _write_detail_sheet(
    sheet: Worksheet,
    result: AnalysisResult,
    plate: PlateData,
    split_report: SplitReport | None,
) -> None:
    """自上而下写四段，段与段之间留一个空行。"""
    row = _write_well_detail(sheet, 1, result, plate)
    row = _write_group_summary(sheet, row + 2, result)
    row = _write_replicate_qc(sheet, row + 2, result)
    row = _write_parameters(sheet, row + 2, result, plate, split_report)
    if result.warnings:
        _write_warnings(sheet, row + 2, result)
    _autofit(sheet, min_width=10, max_width=46)


def _replicate_columns(well: WellRecord | None, fallback: str) -> list[object]:
    """逐孔明细末尾的两列：原样本名、重复编号。未编号的孔编号写 "-"。"""
    if well is None:
        return [fallback, "-"]
    index = well.replicate_index
    return [well.original_sample or well.sample, str(index) if index >= 1 else "-"]


def _write_well_detail(
    sheet: Worksheet, start_row: int, result: AnalysisResult, plate: PlateData
) -> int:
    """逐孔明细：先写进入结果的孔，再补上所有被排除的孔并注明原因。

    末尾两列记录拆分前的样本名和生物学重复编号，这样即使界面上只看得到 CT-2，
    报告里也能一眼对回下机表的 CT。
    """
    headers = [
        "孔位", "基因", "样本", "分组", "Cq", "参与计算",
        "内参均值", "ΔCt", "ΔΔCt", "2^-ΔΔCt", "原样本名", "重复编号",
    ]
    formats = [
        None, None, None, None, NUM_FORMAT, None,
        NUM_FORMAT, NUM_FORMAT, NUM_FORMAT, NUM_FORMAT, None, None,
    ]

    # 用 (孔位, 基因) 而不是带上样本名来对齐：样本名会因拆分开关而变，孔和基因不会
    source_of = {(w.well, w.target): w for w in plate.wells}
    used_keys = {(item.well, item.target) for item in result.well_results}
    rows: list[list[object]] = [
        [
            item.well, item.target, item.sample, item.group, item.cq, "是",
            item.ref_mean, item.dct, item.ddct, item.rq,
        ]
        + _replicate_columns(source_of.get((item.well, item.target)), item.sample)
        for item in result.well_results
    ]
    for well in plate.wells:
        if (well.well, well.target) in used_keys:
            continue
        rows.append(
            [
                well.well, well.target, well.sample, None,
                well.cq if well.valid else well.cq_text,
                f"否（{_exclusion_reason(well, result)}）",
                None, None, None, None,
            ]
            + _replicate_columns(well, well.sample)
        )

    row = _write_section_title(sheet, start_row, "逐孔明细")
    return _write_table(sheet, row, headers, rows, formats)


def _exclusion_reason(well: WellRecord, result: AnalysisResult) -> str:
    """说明这个孔为什么没进结果，可能同时命中多条。"""
    reasons: list[str] = []
    is_reference = well.target in result.reference_targets
    if is_reference:
        reasons.append("内参基因")
    if not well.valid:
        reasons.append("Cq 无效")
    elif not well.included:
        reasons.append("手动剔除")
    if not is_reference and well.usable:
        if well.sample not in result.ref_means:
            reasons.append("该样本缺少内参数据")
        elif well.target not in result.targets:
            reasons.append("该基因在对照组无数据")
    return "；".join(reasons) or "未纳入结果"


def _write_group_summary(
    sheet: Worksheet, start_row: int, result: AnalysisResult
) -> int:
    """分组汇总：每个 (基因, 分组) 的 n / 均值 / 标准差。"""
    headers = ["基因", "分组", "n", "均值", "标准差"]
    formats = [None, None, INT_FORMAT, NUM_FORMAT, NUM_FORMAT]
    rows: list[list[object]] = []
    for target in result.targets:
        for group in result.groups:
            stat = result.stat(target, group)
            if stat is None:
                rows.append([target, group, 0, None, None])
                continue
            rows.append([target, group, stat.n, stat.mean, stat.sd])
    row = _write_section_title(sheet, start_row, "分组汇总")
    return _write_table(sheet, row, headers, rows, formats)


def _write_replicate_qc(sheet: Worksheet, start_row: int, result: AnalysisResult) -> int:
    """复孔一致性：Cq 标准差超阈值的整行标黄。"""
    headers = ["基因", "样本", "n", "Cq 均值", "Cq 标准差", "是否超阈值"]
    formats = [None, None, INT_FORMAT, NUM_FORMAT, NUM_FORMAT, None]
    rows: list[list[object]] = []
    fills: list[PatternFill | None] = []
    for item in result.qc:
        rows.append(
            [
                item.target, item.sample, item.n, item.cq_mean, item.cq_sd,
                "⚠ 偏高" if item.flagged else "正常",
            ]
        )
        fills.append(_WARN_FILL if item.flagged else None)
    row = _write_section_title(
        sheet, start_row, f"复孔一致性（Cq 标准差 > {CQ_SD_WARN} 标黄）"
    )
    return _write_table(sheet, row, headers, rows, formats, row_fills=fills)


def _replicate_mode_text(split_report: SplitReport | None) -> str:
    if split_report is not None and split_report.enabled:
        return "生物学重复配对（同名样本已拆分）"
    return "技术复孔取平均"


def _orientation_text(split_report: SplitReport | None) -> str:
    """孔位排列方向：既写用户选的填板方向，也写实测到的各区块方向分布。"""
    if split_report is None or not split_report.orientations:
        return "未启用（同名复孔按技术复孔取平均）"
    choice = {
        "auto": "自动识别",
        "row": "横向优先（先横后竖）",
        "column": "纵向优先（先竖后横）",
    }.get(split_report.fill_direction, split_report.fill_direction)
    counts: dict[str, int] = {}
    for kind in split_report.orientations.values():
        counts[kind] = counts.get(kind, 0) + 1
    detail = "、".join(
        f"{ORIENTATION_LABELS.get(kind, kind)} {count} 个区块"
        for kind, count in counts.items()
    )
    return f"{choice}；实测 {detail}"


def _split_rows(split_report: SplitReport | None) -> list[list[object]]:
    """复孔处理相关的参数行，附上拆分过程中产生的提示。"""
    rows: list[list[object]] = [
        ["复孔处理模式", _replicate_mode_text(split_report)],
        ["孔位排列方向", _orientation_text(split_report)],
        [
            "已拆分的样本",
            "、".join(split_report.split_samples)
            if split_report is not None and split_report.split_samples
            else "（无）",
        ],
    ]
    if split_report is not None:
        rows.extend(
            [f"拆分提示 {index}", message]
            for index, message in enumerate(split_report.warnings, start=1)
        )
    return rows


def _write_parameters(
    sheet: Worksheet,
    start_row: int,
    result: AnalysisResult,
    plate: PlateData,
    split_report: SplitReport | None = None,
) -> int:
    """分析参数：让报告脱离软件也能复现这次计算。"""
    usable = sum(1 for w in plate.wells if w.usable)
    rows: list[list[object]] = [
        ["源文件名", os.path.basename(plate.file_path)],
        ["数据所在 sheet", plate.sheet_name],
        ["分析时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["内参基因", "、".join(result.reference_targets)],
        ["对照组", result.control_group],
        ["分析基因", "、".join(result.targets)],
        ["分组顺序", " → ".join(result.groups)],
        ["孔数（总数 / 参与计算）", f"{len(plate.wells)} / {usable}"],
        *_split_rows(split_report),
        [
            "内参基线",
            "ref_mean(sample) = 内参基因在该样本内所有复孔 Cq 的算术平均"
            "（多内参先各自求均值再平均）",
        ],
        ["计算公式 ①", "ΔCt_i = Cq_i − ref_mean(sample)　（逐孔计算，目标基因不预先平均）"],
        ["计算公式 ②", "ΔΔCt_i = ΔCt_i − mean(对照组该基因所有孔的 ΔCt)"],
        ["计算公式 ③", "RQ_i = 2^(−ΔΔCt_i)　（每孔一个值）"],
        ["复孔一致性阈值", f"Cq 标准差 > {CQ_SD_WARN} 时提示（按原样本名分桶）"],
        ["软件版本", __version__],
    ]
    row = _write_section_title(sheet, start_row, "分析参数")
    return _write_table(sheet, row, ["项目", "内容"], rows, [None, None])


def _write_warnings(sheet: Worksheet, start_row: int, result: AnalysisResult) -> int:
    """把分析过程中的警告原样列出来。"""
    rows: list[list[object]] = [
        [f"{index}", message] for index, message in enumerate(result.warnings, start=1)
    ]
    row = _write_section_title(sheet, start_row, "分析警告")
    return _write_table(sheet, row, ["#", "警告内容"], rows, [None, None])


# --------------------------------------------------------------------------
# 通用写表工具
# --------------------------------------------------------------------------


def _write_section_title(sheet: Worksheet, row: int, title: str) -> int:
    """写一行加粗小标题，返回下一行行号。"""
    cell = sheet.cell(row=row, column=1, value=title)
    cell.font = _TITLE_FONT
    cell.alignment = _LEFT
    return row + 1


def _write_table(
    sheet: Worksheet,
    start_row: int,
    headers: list[str],
    rows: list[list[object]],
    number_formats: list[str | None],
    row_fills: list[PatternFill | None] | None = None,
) -> int:
    """写「表头 + 数据」，返回最后一行的行号。"""
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=index, value=header)
        cell.font = _BOLD
        cell.alignment = _CENTER
        cell.fill = _HEADER_FILL
        cell.border = _BORDER

    row = start_row
    for offset, values in enumerate(rows):
        row = start_row + 1 + offset
        fill = row_fills[offset] if row_fills else None
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=index, value=value)
            cell.border = _BORDER
            cell.alignment = _CENTER if isinstance(value, (int, float)) else _LEFT
            fmt = number_formats[index - 1] if index - 1 < len(number_formats) else None
            if fmt and isinstance(value, (int, float)):
                cell.number_format = fmt
            if fill is not None:
                cell.fill = fill
    return row


def _cell_width(value: object, number_format: str) -> int:
    """估算单元格显示宽度，中日韩字符按两个字符算。"""
    if isinstance(value, float) and number_format == NUM_FORMAT:
        return len(f"{value:.4f}")
    text = "" if value is None else str(value)
    return sum(
        2 if unicodedata.east_asian_width(char) in ("W", "F") else 1 for char in text
    )


def _autofit(sheet: Worksheet, min_width: int, max_width: int) -> None:
    """按内容估算列宽并夹在 [min_width, max_width] 之间。"""
    widths: dict[int, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            width = _cell_width(cell.value, cell.number_format)
            if width > widths.get(cell.column, 0):
                widths[cell.column] = width
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = max(
            min_width, min(max_width, width + 3)
        )
